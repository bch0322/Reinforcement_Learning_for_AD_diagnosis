import os
import argparse
import setting as st
import setting_2 as fst
from scipy import stats
import nibabel as nib
from scheduler import GradualWarmupScheduler
import torch
import torch.nn as nn
import construct_model
import numpy as np
from torch.backends import cudnn
import utils as ut
# from plot import plot_raw_MMSE
# from plot import plot_age_prediction
# from plot import plot_age_prediction_others
# from plot import plot_age_prediction_cropped_input
# from plot import plot_age_prediction_others_cropped_input
# from plot import generate_heatmap

from data_load import data_load as DL
from data_load import cwk_data_load as cDL
from data_load import jsy_data_load as jDL
from data_load import jacob_data_load as jcDL
from data_load import aal_data_load as aDL
from plot import *
from test import *
from train import *
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import random

from adamp import AdamP
from radam import RAdam
seed = 1234
random.seed(seed)
np.random.seed(seed)
torch.manual_seed(seed)
torch.cuda.manual_seed(seed)
torch.cuda.manual_seed_all(seed)
torch.backends.cudnn.deterministic = True
torch.backends.cudnn.benchmark = False
# torch.backends.cudnn.benchmark = True
# CUDA_LAUNCH_BLOCKING=1

def main(config):

    """ 1. data process """
    if fst.flag_orig_npy == True:
        print('preparation of the numpy')
        if os.path.exists(st.orig_npy_dir) == False :
            os.makedirs(st.orig_npy_dir)

        """ processing """
        if st.list_data_type[st.data_type_num] == 'Density':
            cDL.Prepare_data_GM_AGE_MMSE()
        elif st.list_data_type[st.data_type_num] == 'ADNI_JSY':
            jDL.Prepare_data_1()
        elif st.list_data_type[st.data_type_num] == 'ADNI_JSY_between_1_and_2':
            jDL.Prepare_data_2()
        elif st.list_data_type[st.data_type_num] == 'ADNI_Jacob_256':
            jcDL.Prepare_data_GM_WM_CSF()
        elif 'ADNI_Jacob' in st.list_data_type[st.data_type_num]:
            jcDL.Prepare_data_GM()
        elif 'ADNI_AAL_256' in st.list_data_type[st.data_type_num]:
            aDL.Prepare_data_GM()

    """ 2. fold index processing """
    if fst.flag_fold_index == True:
        print('preparation of the fold index')
        if os.path.exists(st.fold_index_dir) == False:
            os.makedirs(st.fold_index_dir)

        """ save the fold index """
        if st.list_data_type[st.data_type_num] == 'ADNI_JSY_between_1_and_2':
            ut.preparation_fold_index_2(config)  # a1 > a2
            ut.preparation_fold_index_3(config)  # a2 > a1
        else:
            ut.preparation_fold_index(config)

    """ fold selection """
    start_fold = st.start_fold
    end_fold = st.end_fold

    """ workbook """
    list_dir_result = []
    list_wb = []
    list_ws = []
    for i in range(len(st.list_standard_eval_dir)):
        list_dir_result.append(st.dir_to_save_1 + st.list_standard_eval_dir[i])
        ut.make_dir(dir=list_dir_result[i], flag_rm=False)
        out = ut.excel_setting(start_fold=start_fold, end_fold=end_fold, result_dir=list_dir_result[i], f_name='results')
        list_wb.append(out[0])
        list_ws.append(out[1])

    """ fold """
    list_eval_metric = st.list_eval_metric
    metric_avg = [[[] for j in range(len(st.list_eval_metric))] for i in range(len(st.list_standard_eval_dir))]
    for fold in range(start_fold, end_fold+1):
        print("FOLD : {}".format(fold))

        ## TODO : Directory preparation
        print('-' * 10 + 'Directory preparation' + '-' * 10)

        list_dir_save_model_1 = []
        list_dir_save_model_2 = []
        list_dir_confusion_1 = []
        list_dir_heatmap_1 = []
        for i in range(len(st.list_standard_eval_dir)):
            """ dir to save model """
            list_dir_save_model_1.append(
                st.dir_to_save_1 + st.list_standard_eval_dir[i] + '/weights/fold_{}'.format(fold))
            ut.make_dir(dir=list_dir_save_model_1[i], flag_rm=False)

            list_dir_save_model_2.append(
                st.dir_to_save_2 + st.list_standard_eval_dir[i] + '/weights/fold_{}'.format(fold))
            ut.make_dir(dir=list_dir_save_model_2[i], flag_rm=False)

            list_dir_confusion_1.append(
                st.dir_to_save_1 + st.list_standard_eval_dir[i] + '/confusion/fold_{}'.format(fold))
            ut.make_dir(dir=list_dir_confusion_1[i], flag_rm=False)

            list_dir_heatmap_1.append(
                st.dir_to_save_1 + st.list_standard_eval_dir[i] + '/heatmap/fold_{}'.format(fold))
            ut.make_dir(dir=list_dir_heatmap_1[i], flag_rm=False)

        """ dir to save pyplot """
        dir_pyplot_1 = st.dir_to_save_1 + '/pyplot/fold_{}'.format(fold)
        ut.make_dir(dir=dir_pyplot_1, flag_rm=False)


        """ normal classification tasks """
        list_test_result = []
        print('-' * 10 + 'start training' + '-' * 10)
        """ --------------------------------------- """
        """ ------------ classification------------ """
        """ --------------------------------------- """

        """ model construction """
        print('-' * 10 + 'Model construction' + '-' * 10)

        model_1 = construct_model.construct_model(config, flag_model_num=0)
        model_1 = nn.DataParallel(model_1)

        if fst.flag_pretrained ==True:
            dir_to_load = st.dir_preTrain_1
            dir_load_model = dir_to_load + '/weights/fold_{}'.format(fold)

            ##TODO : not exist
            if os.path.exists(dir_load_model) == False :
                print('no directory!!')
            else:
                print('load pretrained')
                model_dir = ut.model_dir_to_load(fold, dir_load_model)
                pretrained_dict = torch.load(model_dir)
                model_dict = model_1.state_dict()

                # for k, v in fst.dict_pretrained_layer.items():
                #     for k_pre, v_pre in pretrained_dict.items():
                #         for k_fine, v_fine in model_dict.items():
                #             if k in k_pre.split('.')[1]:
                #                 if v in k_fine.split('.')[1] :
                #                     if k_fine.split('.')[2:] == k_pre.split('.')[2:] :
                #                         print(k_fine, k_pre)

                if fst.flag_copy_including ==False :
                    dict_weight_to_be_copied = {k_fine : v_pre
                                                for k, v in fst.dict_pretrained_layer.items()
                                                for k_pre, v_pre in pretrained_dict.items()
                                                for k_fine, v_fine in model_dict.items()
                                                if k == k_pre.split('.')[1]
                                                if v == k_fine.split('.')[1]
                                                if k_fine.split('.')[2:] == k_pre.split('.')[2:]
                                                }
                else:
                    dict_weight_to_be_copied = {k_fine : v_pre
                                                for k, v in fst.dict_pretrained_layer.items()
                                                for k_pre, v_pre in pretrained_dict.items()
                                                for k_fine, v_fine in model_dict.items()
                                                if k in k_pre.split('.')[1]
                                                if v in k_fine.split('.')[1]
                                                if k_fine.split('.')[2:] == k_pre.split('.')[2:]
                                                }

                model_dict.update(dict_weight_to_be_copied)
                model_1.load_state_dict(model_dict)
                ut.dfs_freeze_with_key(model_1, fst.list_freeze_layer, requires_grad=False)

        """ optimizer """
        optimizer_1 = torch.optim.Adam(model_1.parameters(), lr=st.hyperParam_s1.lr, betas=(0.9, 0.999), eps=1e-8, weight_decay=st.hyperParam_s1.weight_decay)
        # optimizer_1 = torch.optim.SGD(model_1.parameters(), lr=st.hyperParam_s1.lr, momentum=0.9, weight_decay=st.hyperParam_s1.weight_decay)
        # optimizer_1 = AdamP(model_1.parameters(), lr=hyperParam_s1.lr, betas=(0.9, 0.999), weight_decay=st.hyperParam_s1.weight_decay)
        # optimizer_1 = RAdam(model_1.parameters(), lr=hyperParam_s1.lr, betas=(0.9, 0.999), weight_decay=st.hyperParam_s1.weight_decay)

        """ expo """
        if fst.flag_scheduler_expo == True:
            scheduler_1 = torch.optim.lr_scheduler.StepLR(optimizer_1, step_size=st.hyperParam_s1.step_size, gamma=st.hyperParam_s1.LR_decay_rate, last_epoch=-1)

        """ cosine annealing """
        if fst.flag_scheduler_cosine == True:
            scheduler_cosine_1 = torch.optim.lr_scheduler.CosineAnnealingLR(optimizer_1, st.hyperParam_s1.epoch)
            scheduler_1 = GradualWarmupScheduler(optimizer_1, multiplier=1, total_epoch=5, after_scheduler=scheduler_cosine_1)

        """ etc """
        # scheduler_cosine_restart = torch.optim.lr_scheduler.CosineAnnealingWarmRestarts(optimizer=optimizer, T_0=50)
        # scheduler = GradualWarmupScheduler(optimizer, multiplier=1, total_epoch=5, after_scheduler=scheduler_cosine_restart)
        # scheduler = torch.optim.lr_scheduler.CosineAnnealingWarmRestarts(optimizer=optimizer, T_0=50)
        # scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(optimizer=optimizer, mode='min', factor=0.2, patience=10)

        if fst.flag_stage_1_RL == True:
            model_2 = construct_model.construct_model(config, flag_model_num=1)
            model_2 = nn.DataParallel(model_2)
            optimizer_2 = torch.optim.Adam(model_2.parameters(), lr=st.hyperParam_s2.lr, betas=(0.9, 0.999), eps=1e-8, weight_decay=st.hyperParam_s2.weight_decay)

            """ expo """
            if fst.flag_scheduler_expo == True:
                scheduler_2 = torch.optim.lr_scheduler.StepLR(optimizer_2, step_size=st.hyperParam_s2.step_size,
                                                              gamma=st.hyperParam_s2.LR_decay_rate, last_epoch=-1)

            """ cosine annealing """
            if fst.flag_scheduler_cosine == True:
                scheduler_cosine_2 = torch.optim.lr_scheduler.CosineAnnealingLR(optimizer_2, st.hyperParam_s2.epoch)
                scheduler_2 = GradualWarmupScheduler(optimizer_2, multiplier=1, total_epoch=5, after_scheduler=scheduler_cosine_2)


        """ data loader """
        print('-' * 10 + 'data loader' + '-' * 10)


        train_loader = DL.convert_Dloader_3(fold, list_class=st.list_class_for_train, flag_tr_val_te='train', batch_size=st.hyperParam_s1.batch_size, num_workers=0, shuffle=True, drop_last=True)
        val_loader = DL.convert_Dloader_3(fold, list_class=st.list_class_for_test,  flag_tr_val_te='val', batch_size=st.hyperParam_s1.v_batch_size, num_workers=0, shuffle=False, drop_last=False)
        test_loader = DL.convert_Dloader_3(fold, list_class=st.list_class_for_test,  flag_tr_val_te='test', batch_size=st.hyperParam_s1.v_batch_size, num_workers=0, shuffle=False, drop_last=False)
        dict_data_loader = {'train' : train_loader,
                            'val' : val_loader,
                            'test' : test_loader}
        
        if fst.flag_stage_1 == True:
            train.train(config, fold, model_1, dict_data_loader, optimizer_1, scheduler_1, st.hyperParam_s1, list_dir_save_model_1, list_dir_heatmap_1,  dir_pyplot_1, Validation=True, Test_flag=True)

        elif fst.flag_stage_1_RL == True:
            train_2.train(config, fold, model_1, model_2, dict_data_loader, optimizer_1, optimizer_2, scheduler_1, scheduler_2, st.hyperParam_s1, list_dir_save_model_1, list_dir_save_model_2, list_dir_heatmap_1, dir_pyplot_1, Validation=True, Test_flag=True)



        """ test classification model """
        if fst.flag_stage_1 == True:
            for i_tmp in range(len(st.list_standard_eval_dir)):
                dict_test_output = test.test(config, fold, model_1, dict_data_loader['test'], st.hyperParam_s1, list_dir_save_model_1[i_tmp], list_dir_heatmap_1[i_tmp], list_dir_confusion_1[i_tmp])
                list_test_result.append(dict_test_output)
        elif fst.flag_stage_1_RL == True:
            for i_tmp in range(len(st.list_standard_eval_dir)):
                dict_test_output = test_2.test(config, fold, model_1, model_2, dict_data_loader['test'], st.hyperParam_s1, list_dir_save_model_1[i_tmp], list_dir_save_model_2[i_tmp], list_dir_heatmap_1[i_tmp], list_dir_confusion_1[i_tmp])
                list_test_result.append(dict_test_output)

        """ fill out the results on the excel sheet """
        for i_standard in range(len(st.list_standard_eval_dir)):
            for i in range(len(list_eval_metric)):
                if list_eval_metric[i] in list_test_result[i_standard]:
                    list_ws[i_standard].cell(row=2 + i + st.push_start_row, column=fold + 1, value="%.4f" % (list_test_result[i_standard][list_eval_metric[i]]))
                    metric_avg[i_standard][i].append(list_test_result[i_standard][list_eval_metric[i]])

            for i in range(len(list_eval_metric)):
                if metric_avg[i_standard][i]:
                    avg = round(np.mean(metric_avg[i_standard][i]), 4)
                    std = round(np.std(metric_avg[i_standard][i]), 4)
                    tmp = "%.4f \u00B1 %.4f" % (avg, std)
                    list_ws[i_standard].cell(row=2 + st.push_start_row + i, column=end_fold + 2, value=tmp)

            n_row = list_ws[i_standard].max_row
            n_col = list_ws[i_standard].max_column
            for i_row in range(1, n_row + 1):
                for i_col in range(1, n_col + 1):
                    ca1 = list_ws[i_standard].cell(row=i_row, column=i_col)
                    ca1.alignment = Alignment(horizontal='center', vertical='center')
            list_wb[i_standard].save(list_dir_result[i_standard] + "/results.xlsx")
            list_wb[i_standard].close()

        del model_1, train_loader, test_loader, optimizer_1
        print("finished_stage_1")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--kfold', type=int, default=st.kfold)
    parser.add_argument('--num_classes', type=int, default=st.num_class)

    parser.add_argument('--sagital', type=int, default=st.x_size)
    parser.add_argument('--coronal', type=int, default=st.y_size)
    parser.add_argument('--axial', type=int, default=st.z_size)
    parser.add_argument('--modality', type=int, default=st.num_modality)

    parser.add_argument('--lr', type=float, default=st.hyperParam_s1.lr)
    parser.add_argument('--batch_size',type=int, default=st.hyperParam_s1.batch_size)
    parser.add_argument('--v_batch_size', type=int, default=st.hyperParam_s1.v_batch_size)
    parser.add_argument('--num_epochs', type=int, default=st.hyperParam_s1.epoch)
    parser.add_argument('--selected_model', type=str, default=st.model_name)

    config = parser.parse_args()
    main(config)

